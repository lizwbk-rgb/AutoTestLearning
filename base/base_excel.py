# -*- coding: utf-8 -*-
"""
兼容 Python 3.8 的 .xlsx 读写工具
读取：openpyxl -> list[dict]
写入：openpyxl <- list[dict]
"""
from typing import List, Dict, Optional          # 兼容 3.8 泛型
from openpyxl import load_workbook, Workbook     # 读写 xlsx


# -------------------- 读取 --------------------
class ExcelRead:
    """
    读取 *.xlsx 文件
    用法：
        excel = ExcelRead("demo.xlsx", "Sheet1")
        data  = excel.dict_data()   # List[Dict]
    """

    def __init__(self, excel_path: str, sheet_name: str = "Sheet1"):
        """
        加载工作簿
        :param excel_path: 文件路径
        :param sheet_name: 工作表名
        """
        self.wb = load_workbook(excel_path)
        self.ws = self.wb[sheet_name]
        self.row_num = self.ws.max_row
        self.col_num = self.ws.max_column
        # 第一行当表头
        self.headers: List[str] = [
            self.ws.cell(row=1, column=j).value for j in range(1, self.col_num + 1)
        ]

    def dict_data(self) -> List[Dict[str, object]]:
        """转为 List[Dict] 返回"""
        if self.row_num <= 1:
            print("总行数 <= 1，无数据")
            return []

        ret: List[Dict[str, object]] = []
        for row in range(2, self.row_num + 1):
            ret.append(
                {self.headers[col - 1]: self.ws.cell(row=row, column=col).value
                 for col in range(1, self.col_num + 1)}
            )
        return ret

    def get_row_info(self, row: int) -> Optional[Dict[str, object]]:
        """取第 row 行（从 2 开始）数据"""
        return None if self.row_num <= 1 else self.dict_data()[row - 2]

    def get_col_info(self, col_name: str) -> List[object]:
        """取整列数据"""
        return [row[col_name] for row in self.dict_data()]

    def get_cell_info(self, row: int, col_name: str) -> object:
        """取单个单元格"""
        return self.get_row_info(row)[col_name]


# -------------------- 写入 --------------------
class ExcelWrite:
    """
    写入 *.xlsx
    用法：
        w = ExcelWrite("Sheet1")
        w.write_excel(data, "result.xlsx")
    """

    def __init__(self, sheet_name: str = "Sheet1"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

    def set_header(self, data: List[Dict[str, object]]) -> None:
        """根据字典 key 写表头（第一行）"""
        if not data:
            return
        for col, key in enumerate(data[0].keys(), 1):
            self.ws.cell(row=1, column=col, value=key)

    def write_excel(self, data: List[Dict[str, object]], save_path: str) -> None:
        """
        将 List[Dict] 数据完整写入 .xlsx 并保存
        :param data: 待写入数据，外层 List 每个元素为一行 Dict
        :param save_path: 输出文件路径（须以 .xlsx 结尾）
        """
        self.set_header(data)  # 写入表头（第一行）
        for row_idx, row_dict in enumerate(data, start=2):  # 从第二行开始写数据
            for col_idx, value in enumerate(row_dict.values(), start=1):
                self.ws.cell(row=row_idx, column=col_idx, value=value)
        self.wb.save(save_path)  # 落盘
        print(f"已保存 -> {save_path}")


# -------------------- 自测 --------------------
if __name__ == "__main__":
    # 1. 读示例
    read_path = r"../data/data_driver/excel_driver/project_auto_test/test.xlsx"
    reader = ExcelRead(read_path, "Sheet1")
    print("读取结果：", reader.dict_data())

    # 2. 写示例
    demo: List[Dict[str, object]] = [
        {"ID": 1, "用户名": "admin", "密码": "123456"},
        {"ID": 2, "用户名": "guest", "密码": "abcde"}
    ]
    # ExcelWrite("Sheet1").write_excel(demo, "output_demo.xlsx")